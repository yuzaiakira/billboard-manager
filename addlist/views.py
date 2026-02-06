from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import ListsModel
from .forms import PdfExportFieldForm, ExcelExportFieldForm
from billboard.utils import billboard_bool_value
# Create your views here.


class AddToList(LoginRequiredMixin, View):
    http_method_names = ['get']

    def get(self, request, pk):
        item, created = ListsModel.objects.get_or_create(
            user=request.user,
            billboard_id=pk
        )
        item.save()
        return JsonResponse({
            'id': pk,
            'in_list': not created,
            'added': created
        })


class RemoveFromList(AddToList):
    def get(self, request, pk):
        try:
            item = ListsModel.objects.get(user=request.user, billboard_id=pk)
            item.delete()
            deleted = True
        except ListsModel.DoesNotExist:
            deleted = False

        return JsonResponse({
            'id': pk,
            'deleted': deleted
        })


class WatchList(LoginRequiredMixin, View):
    http_method_names = ['get']
    model = ListsModel
    template = 'template/list/Billboard_watch_list.html'
    context = dict()
    context['pdf_form'] = PdfExportFieldForm()
    context['excel_form'] = ExcelExportFieldForm()

    def get(self, request):
        self.context['object_list'] = self.model.objects.filter(user=request.user)
        self.context['title'] = "لیست بیلبورد های انتخاب شده"
        # TODO: add only
        return render(request, self.template, self.context)


class RemoveList(LoginRequiredMixin, View):
    http_method_names = ['get']
    model = ListsModel

    def get(self, request):
        user = request.user
        list_model = self.model.objects.filter(user=user)
        list_model.delete()
        return redirect('WatchList')


class PrintPDF(WatchList):
    template = 'template/list/Print_PDF.html'

    def get(self, request, *args, **kwargs):
        if request.method == 'GET':
            self.context['pdf_form'] = PdfExportFieldForm(request.GET)
            if self.context['pdf_form'].is_valid():
                print(self.context['pdf_form'].cleaned_data['billboard_pic'])
                print(self.context['pdf_form'].cleaned_data)

        return super().get(request, *args, **kwargs)


class ExportExcel(LoginRequiredMixin, View):
    model = ListsModel

    # Map ExcelExportFieldForm field names to (persian_header, variable_header, value_key)
    _EXCEL_COLUMNS = {
        'id_code': ('کد بیلبورد', 'id_code', 'id'),
        'city': ('شهر', 'city', 'city'),
        'name': ('عنوان بیلبورد', 'name', 'name'),
        'address': ('آدرس بیلبورد', 'address', 'address'),
        'description': ('توضیحات بیلبورد', 'description', 'description'),
        'has_power': ('روشنایی بیلبورد', 'has_power', 'has_power'),
        'size': ('طول', 'billboard_length', 'billboard_length'),  # size adds two columns
        'size_width': ('عرض', 'billboard_width', 'billboard_width'),
        'reservation_date': ('تاریخ رزرو بیلبورد', 'reservation_date', 'reservation_date'),
        'price': ('قیمت بیلبورد', 'price', 'price'),
    }

    def get(self, request, *args, **kwargs):
        form = ExcelExportFieldForm(request.GET)
        initial_data = {f: form.fields[f].initial for f in form.fields}
        if form.is_valid():
            data = form.cleaned_data if request.GET else initial_data
            use_persian_headers = data.get('style', True)
            columns = self._build_columns_from_form(data, use_persian_headers)
        else:
            use_persian_headers = initial_data.get('style', True)
            columns = self._build_columns_from_form(initial_data, use_persian_headers)
        if not columns:
            columns = self._default_columns(use_persian_headers)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="billboard-list.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "billboard list"

        headers = [col[0] for col in columns]  # col = (header, value_key)
        ws.append(headers)

        items = self.model.objects.filter(user=request.user)
        for item in items:
            billboard = item.billboard
            row = []
            for _, value_key in columns:
                if value_key == 'id':
                    row.append(str(billboard.id))
                elif value_key == 'city':
                    row.append(str(billboard.city))
                elif value_key == 'name':
                    row.append(billboard.name)
                elif value_key == 'address':
                    row.append(billboard.address)
                elif value_key == 'description':
                    row.append(billboard.description or '')
                elif value_key == 'has_power':
                    row.append(billboard_bool_value(billboard.has_power))
                elif value_key == 'billboard_length':
                    row.append(billboard.billboard_length or '')
                elif value_key == 'billboard_width':
                    row.append(billboard.billboard_width or '')
                elif value_key == 'reservation_date':
                    d = billboard.reservation_date
                    row.append(f"{d.year}/{d.month:02d}/{d.day:02d}" if d and hasattr(d, 'year') else (str(d) if d else ''))
                elif value_key == 'price':
                    row.append(billboard.price or '')
                else:
                    row.append('')
            ws.append(row)

        if use_persian_headers:
            self._apply_persian_style(ws, len(columns), len(items))

        wb.save(response)
        return response

    def _apply_persian_style(self, ws, num_columns, num_data_rows):
        """Apply RTL and professional styling when ظاهر فارسی is enabled."""
        ws.sheet_view.rightToLeft = True

        # Style constants
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
        cell_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        data_fill_light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        data_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(name="Tahoma", size=10, color="000000")
        alignment_rtl = Alignment(horizontal="right", vertical="center", wrap_text=True)

        # Header row (row 1)
        for col_idx in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data rows
        for row_idx in range(2, num_data_rows + 2):
            row_fill = data_fill_light if row_idx % 2 == 0 else data_fill_white
            for col_idx in range(1, num_columns + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.font = data_font
                cell.border = cell_border
                cell.alignment = alignment_rtl

        # Column widths: wider than content for readability (min 14, max 40)
        for col_idx in range(1, num_columns + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, min(num_data_rows + 2, 102) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    try:
                        max_length = max(max_length, len(str(val)))
                    except (TypeError, ValueError):
                        pass
            adjusted = min(max(max_length + 2, 14), 40)
            ws.column_dimensions[col_letter].width = adjusted

    def _build_columns_from_form(self, cleaned_data, use_persian_headers=True):
        """Build (header, value_key) list from form. Header is Persian or variable name based on style."""
        columns = []
        for field_name, (persian_header, variable_header, value_key) in self._EXCEL_COLUMNS.items():
            if field_name == 'size_width':
                if cleaned_data.get('size'):
                    header = persian_header if use_persian_headers else variable_header
                    columns.append((header, value_key))
                continue
            if cleaned_data.get(field_name):
                header = persian_header if use_persian_headers else variable_header
                columns.append((header, value_key))
        return columns

    def _default_columns(self, use_persian_headers=True):
        """Default columns when all fields are turned off (aligned with the form's initial state)."""
        if use_persian_headers:
            return [
                ('کد بیلبورد', 'id'),
                ('شهر', 'city'),
                ('آدرس بیلبورد', 'address'),
                ('توضیحات بیلبورد', 'description'),
                ('روشنایی بیلبورد', 'has_power'),
                ('طول', 'billboard_length'),
                ('عرض', 'billboard_width'),
                ('قیمت بیلبورد', 'price'),
                ('تاریخ رزرو بیلبورد', 'reservation_date'),
            ]
        return [
            ('id_code', 'id'),
            ('city', 'city'),
            ('address', 'address'),
            ('description', 'description'),
            ('has_power', 'has_power'),
            ('billboard_length', 'billboard_length'),
            ('billboard_width', 'billboard_width'),
            ('price', 'price'),
            ('reservation_date', 'reservation_date'),
        ]
